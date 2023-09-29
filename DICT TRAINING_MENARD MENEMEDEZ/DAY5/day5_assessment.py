from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, messagebox


root = Tk()
root.geometry('1100x500')
root.configure(bg='lightblue')
root.title('CRUD EME')


excel_connection = Workbook()
excel_connection = load_workbook('register.xlsx')
excel_opr = excel_connection.active


name_label = Label(root, text='Name : ', font='Arial 16', bg='lightblue')
name_label.grid(row=0, column=0, pady=5)

name_entry = Entry(root, width=20, font='Arial 16')
name_entry.grid(row=0, column=1, pady=5)


program_label = Label(root, text='Program Enrolled : ', font='Arial 16', bg='lightblue')
program_label.grid(row=1, column=0, rowspan=5, pady=5)


program_selected = StringVar(root, 'BSIT')

BSIT = Radiobutton(root, text='BS in Information Technology', variable=program_selected, value='BSIT', bg='lightblue')
BSCS = Radiobutton(root, text='BS in Computer Science', variable=program_selected, value='BSCS', bg='lightblue')
BSCE = Radiobutton(root, text='BS in Computer Engineering', variable=program_selected, value='BSCE', bg='lightblue')
BSIS = Radiobutton(root, text='BS in Information System', variable=program_selected, value='BSIS', bg='lightblue')


BSIT.grid(row=2, column=1, pady=2, sticky='w')
BSCS.grid(row=3, column=1, pady=2, sticky='w')
BSCE.grid(row=4, column=1, pady=2, sticky='w')
BSIS.grid(row=5, column=1, pady=2, sticky='w')


address_label = Label(root, text='Residence : ', font='Arial 16', bg='lightblue')
address_label.grid(row=6, column=0, pady=4)

address = Text(root, width=30, height=3)
address.grid(row=6, column=1, pady=4, sticky='w')


submit = Button(root, text='Submit Record', width=12, height=2, font='Arial 10', command=lambda:submit_data())
submit.grid(row=7, column=1, pady=5)

search_btn = Button(root, text='Search', width=12, height=2, font='Arial 10', command=lambda:search_data())
search_btn.grid(row=8, column=1, pady=5)

delete_btn = Button(root, text='Delete', width=12, height=2, font='Arial 10', command=lambda:delete_data())
delete_btn.grid(row=9, column=1, pady=5)

update_btn = Button(root, text='Update', width=12, height=2, font='Arial 10')
update_btn.grid(row=10, column=1, pady=5)

tree = ttk.Treeview(root, columns=('NAMES', 'PROGRAM', 'ADDRESS'), show='headings')
tree.heading('NAMES', text='NAMES')
tree.heading('PROGRAM', text='PROGRAM')
tree.heading('ADDRESS', text='ADDRESS')

tree.grid(row=0, column=3, rowspan=7, padx=10)

def insert_data():
    for row in excel_opr.iter_rows():
        tree.insert('','end',values=(row[0].value, row[1].value, row[2].value))



def search_data():
    nameTxt = name_entry.get()
    found = False

    i = 1
    for row in excel_opr.iter_rows():
        if nameTxt in row[0].value:
            found = True
            break
        i = i + 1

    if found:
        # name_entry.insert(excel_opr[f'A{i}'])
        program_value = excel_opr[f'B{i}']
        if program_value == 'BSIT':
            BSIT.select(True)
        elif program_value == 'BSCS':
            BSCS.select(True)
        elif program_value == 'BSCE':
            BSCE.select(True)
        elif program_value == 'BSIS':
            BSIS.select(True)

        msg = f'We found a record related to your input name'
        messagebox.showinfo('Found', msg)
    else:
        msg = f'Data not found'
        messagebox.showinfo('Not Found', msg)

def get_updated_data():
    tree.delete(*tree.get_children())
    for row in excel_opr.iter_rows():
        tree.insert('','end',values=(row[0].value, row[1].value, row[2].value))




def delete_data():
    nameTxt = name_entry.get()
    found = False
    i = 1
    for row in excel_opr.iter_rows():
        if nameTxt in row[0].value:
            found = True
            break
        i = i + 1
    
    if found:
        excel_opr.delete_rows(i)
        messagebox.showerror('DATA DELETED', 'DATA REMOVED FROM THE DB')
        excel_connection.save('register.xlsx')
    else:
        messagebox.showerror('NO DATA FOUND', 'SORRY, WE COULDN`T FIND THE NAME!')

    get_updated_data()


def submit_data():
    nameTxt = name_entry.get()
    programTxt = program_selected.get()
    addressTxt = address.get('1.0', 'end-1c')

    last_row = excel_opr.max_row

    excel_opr[f'A{last_row +1}'] = nameTxt
    excel_opr[f'B{last_row + 1}'] = programTxt
    excel_opr[f'C{last_row + 1}'] = addressTxt

    excel_connection.save('register.xlsx')

    msg = f'Student {nameTxt}, enrolled in {programTxt},.\nCurrently residing in {addressTxt}'
    messagebox.showinfo('Record', msg)
    tree.insert('','end',values=(nameTxt, programTxt, addressTxt))

insert_data()
root.mainloop()